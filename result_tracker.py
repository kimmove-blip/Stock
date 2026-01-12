"""
결과 추적 모듈
전날 선정 종목의 다음날 실적을 기록 + 투자 시뮬레이션
"""

import os
import glob
from datetime import datetime, timedelta
import pandas as pd
import FinanceDataReader as fdr
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from config import OUTPUT_DIR


def get_previous_result_file():
    """전날 결과 파일 찾기 (오늘 파일 제외)"""
    pattern = os.path.join(OUTPUT_DIR, "top100_*.xlsx")
    files = glob.glob(pattern)

    if not files:
        return None

    today_str = datetime.now().strftime("%Y%m%d")

    # 오늘 파일 제외하고 가장 최근 파일
    files = [f for f in files if today_str not in f]

    if not files:
        return None

    files.sort(key=os.path.getmtime, reverse=True)
    return files[0]


def get_stock_ohlc(code, days=7):
    """종목의 OHLC 데이터 조회"""
    today = datetime.now()
    start_date = (today - timedelta(days=days)).strftime("%Y-%m-%d")
    end_date = today.strftime("%Y-%m-%d")

    try:
        df = fdr.DataReader(code, start_date, end_date)
        if not df.empty:
            return df
    except:
        pass
    return None


def calculate_investment(open_price, close_price, investment_per_stock=100000):
    """
    투자 시뮬레이션 계산
    - 시초가에 매수, 종가에 매도
    - 종목당 10만원 투자 (1주당 10만원 넘으면 1주만)

    Returns:
        dict: 매입주수, 투자금, 회수금, 수익금, 수익률
    """
    if open_price is None or close_price is None or open_price <= 0:
        return None

    # 주식 수 계산
    if open_price >= investment_per_stock:
        shares = 1  # 1주만 매입
    else:
        shares = int(investment_per_stock // open_price)

    if shares == 0:
        shares = 1

    # 투자금 = 시초가 * 주식수
    investment = int(open_price * shares)

    # 회수금 = 종가 * 주식수
    returns = int(close_price * shares)

    # 수익금
    profit = returns - investment

    # 수익률
    profit_rate = round((profit / investment) * 100, 2) if investment > 0 else 0

    return {
        'shares': shares,
        'investment': investment,
        'returns': returns,
        'profit': profit,
        'profit_rate': profit_rate
    }


def get_yesterday_results(prev_excel_path):
    """
    전날 선정 종목의 오늘 실적 조회
    시초가 매수 → 종가 매도 기준
    """
    if not prev_excel_path or not os.path.exists(prev_excel_path):
        return None, None

    # 전날 데이터 읽기
    df = pd.read_excel(prev_excel_path)

    # 상위 30개만
    df = df.head(30)

    codes = df['종목코드'].astype(str).str.zfill(6).tolist()

    results = []
    total_investment = 0
    total_returns = 0
    trade_date = None

    for _, row in df.iterrows():
        code = str(row['종목코드']).zfill(6)
        name = row['종목명']
        prev_score = row.get('종합점수', 0)
        prev_close = row.get('현재가', 0)
        target_price = row.get('목표가', 0)

        # 오늘 OHLC 조회
        ohlc = get_stock_ohlc(code, days=3)

        if ohlc is not None and len(ohlc) >= 1:
            latest = ohlc.iloc[-1]
            today_open = int(latest['Open'])
            today_close = int(latest['Close'])
            today_high = int(latest['High'])
            today_low = int(latest['Low'])

            if trade_date is None:
                trade_date = ohlc.index[-1].strftime("%Y-%m-%d")

            # 투자 시뮬레이션
            inv_result = calculate_investment(today_open, today_close)

            if inv_result:
                total_investment += inv_result['investment']
                total_returns += inv_result['returns']

                # 수익률 계산 (전일 종가 대비)
                return_rate = round((today_close - prev_close) / prev_close * 100, 2) if prev_close > 0 else 0

                # 결과 판정
                if return_rate > 0:
                    result_text = "✅ 수익"
                elif return_rate < 0:
                    result_text = "❌ 손실"
                else:
                    result_text = "➖ 보합"

                # 목표가 달성 여부
                target_reached = "✅" if today_high >= target_price else "❌"

                results.append({
                    '순위': len(results) + 1,
                    '종목코드': code,
                    '종목명': name,
                    '선정일점수': prev_score,
                    '선정일종가': prev_close,
                    '목표가': target_price,
                    '금일시가': today_open,
                    '금일종가': today_close,
                    '수익률(%)': return_rate,
                    '결과': result_text,
                    '목표달성': target_reached,
                    '매입주수': inv_result['shares'],
                    '투자금': inv_result['investment'],
                    '회수금': inv_result['returns'],
                    '수익금': inv_result['profit'],
                })
        else:
            results.append({
                '순위': len(results) + 1,
                '종목코드': code,
                '종목명': name,
                '선정일점수': prev_score,
                '선정일종가': prev_close,
                '목표가': target_price,
                '금일시가': '-',
                '금일종가': '-',
                '수익률(%)': '-',
                '결과': '조회실패',
                '목표달성': '-',
                '매입주수': '-',
                '투자금': '-',
                '회수금': '-',
                '수익금': '-',
            })

    # 요약 통계
    summary = {
        'trade_date': trade_date or datetime.now().strftime("%Y-%m-%d"),
        'total_stocks': len(results),
        'total_investment': total_investment,
        'total_returns': total_returns,
        'total_profit': total_returns - total_investment,
        'total_profit_rate': round((total_returns - total_investment) / total_investment * 100, 2) if total_investment > 0 else 0,
    }

    # 성공/실패 통계
    valid_results = [r for r in results if isinstance(r.get('수익률(%)'), (int, float))]
    if valid_results:
        profits = [r['수익률(%)'] for r in valid_results]
        summary['avg_return'] = round(sum(profits) / len(profits), 2)
        summary['max_return'] = max(profits)
        summary['min_return'] = min(profits)
        summary['success_count'] = len([r for r in valid_results if r['수익률(%)'] > 0])
        summary['fail_count'] = len([r for r in valid_results if r['수익률(%)'] < 0])
        summary['target_reached_count'] = len([r for r in results if r.get('목표달성') == '✅'])

        # 최고/최저 수익 종목
        best = max(valid_results, key=lambda x: x['수익률(%)'])
        worst = min(valid_results, key=lambda x: x['수익률(%)'])
        summary['best_stock'] = f"{best['종목명']} ({best['수익률(%)']}%)"
        summary['worst_stock'] = f"{worst['종목명']} ({worst['수익률(%)']}%)"

    return pd.DataFrame(results), summary


def style_excel_workbook(wb):
    """엑셀 워크북 스타일링"""
    # 스타일 정의
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cell_alignment = Alignment(horizontal="center", vertical="center")
    number_alignment = Alignment(horizontal="right", vertical="center")

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    for ws in wb.worksheets:
        # 헤더 스타일
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 데이터 셀 스타일
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = cell_alignment

                # 숫자 포맷
                if isinstance(cell.value, (int, float)):
                    if '투자금' in str(ws.cell(1, cell.column).value) or \
                       '회수금' in str(ws.cell(1, cell.column).value) or \
                       '수익금' in str(ws.cell(1, cell.column).value) or \
                       '현재가' in str(ws.cell(1, cell.column).value) or \
                       '종가' in str(ws.cell(1, cell.column).value) or \
                       '시가' in str(ws.cell(1, cell.column).value) or \
                       '목표가' in str(ws.cell(1, cell.column).value):
                        cell.number_format = '#,##0'
                    elif '%' in str(ws.cell(1, cell.column).value):
                        cell.number_format = '0.00'

                # 결과 컬럼 색상
                if '결과' in str(ws.cell(1, cell.column).value):
                    if '수익' in str(cell.value):
                        cell.fill = success_fill
                    elif '손실' in str(cell.value):
                        cell.fill = fail_fill

        # 컬럼 너비 자동 조정
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width

    return wb


def create_two_sheet_excel(today_results, yesterday_df, yesterday_summary, output_path):
    """
    2개 시트가 있는 엑셀 파일 생성
    - 시트1: 내일의 관심종목
    - 시트2: 전일 관심종목 결과
    """
    from config import get_signal_kr

    # 시트1: 내일의 관심종목 (상위 30개)
    top30 = today_results[:30]

    sheet1_rows = []
    for i, r in enumerate(top30, 1):
        signals = r.get("signals", [])
        patterns = r.get("patterns", [])

        signals_kr = [get_signal_kr(s) for s in (signals or [])[:3]]

        score = r["score"]
        close_price = int(r.get("close", 0))
        if score >= 80:
            target_mul = 1.20
        elif score >= 60:
            target_mul = 1.10
        elif score >= 40:
            target_mul = 1.05
        else:
            target_mul = 1.0
        target_price = int(close_price * target_mul)

        sheet1_rows.append({
            "순위": i,
            "종목코드": r["code"],
            "종목명": r["name"],
            "시장": r["market"],
            "종합점수": r["score"],
            "현재가": close_price,
            "목표가": target_price,
            "기대수익률(%)": round((target_mul - 1) * 100, 1),
            "등락률(%)": round(r.get("change_pct", 0), 2),
            "거래량": int(r.get("volume", 0)),
            "주요신호": " | ".join(signals_kr) if signals_kr else "-",
            "캔들패턴": " | ".join(patterns) if patterns else "-",
        })

    df_sheet1 = pd.DataFrame(sheet1_rows)

    # 엑셀 저장
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_sheet1.to_excel(writer, sheet_name='내일의 관심종목', index=False)

        if yesterday_df is not None and len(yesterday_df) > 0:
            yesterday_df.to_excel(writer, sheet_name='전일 관심종목 결과', index=False)

            # 요약 정보 추가
            ws = writer.sheets['전일 관심종목 결과']

            # 빈 행 추가 후 요약
            start_row = len(yesterday_df) + 3

            summary_data = [
                ['', ''],
                ['[ 투자 시뮬레이션 결과 ]', ''],
                ['기준일', yesterday_summary.get('trade_date', '-')],
                ['분석 종목 수', f"{yesterday_summary.get('total_stocks', 0)}개"],
                ['총 투자금', f"{yesterday_summary.get('total_investment', 0):,}원"],
                ['총 회수금', f"{yesterday_summary.get('total_returns', 0):,}원"],
                ['총 수익금', f"{yesterday_summary.get('total_profit', 0):,}원"],
                ['총 수익률', f"{yesterday_summary.get('total_profit_rate', 0)}%"],
                ['', ''],
                ['[ 성과 분석 ]', ''],
                ['평균 수익률', f"{yesterday_summary.get('avg_return', 0)}%"],
                ['최고 수익률', f"{yesterday_summary.get('max_return', 0)}%"],
                ['최저 수익률', f"{yesterday_summary.get('min_return', 0)}%"],
                ['수익 종목', f"{yesterday_summary.get('success_count', 0)}개"],
                ['손실 종목', f"{yesterday_summary.get('fail_count', 0)}개"],
                ['성공률', f"{round(yesterday_summary.get('success_count', 0) / yesterday_summary.get('total_stocks', 1) * 100, 1)}%"],
                ['목표가 달성', f"{yesterday_summary.get('target_reached_count', 0)}개"],
                ['', ''],
                ['최고 수익 종목', yesterday_summary.get('best_stock', '-')],
                ['최저 수익 종목', yesterday_summary.get('worst_stock', '-')],
            ]

            for i, row_data in enumerate(summary_data):
                for j, value in enumerate(row_data):
                    ws.cell(row=start_row + i, column=len(yesterday_df.columns) + 2 + j, value=value)

    # 스타일 적용
    wb = load_workbook(output_path)
    wb = style_excel_workbook(wb)
    wb.save(output_path)

    return output_path


def update_with_next_day_results(excel_path=None):
    """
    전날 결과 추적 및 출력
    (기존 호환성 유지용)
    """
    prev_file = get_previous_result_file()

    if prev_file is None:
        print("[추적] 전날 파일을 찾을 수 없습니다.")
        return None

    print(f"[추적] 전날 파일: {prev_file}")

    # 전날 결과 조회
    yesterday_df, summary = get_yesterday_results(prev_file)

    if yesterday_df is None:
        print("[추적] 결과 조회 실패")
        return None

    # 통계 출력
    print(f"\n[추적] 결과 요약 ({summary.get('trade_date', '-')}):")
    print(f"    - 총 투자금: {summary.get('total_investment', 0):,}원")
    print(f"    - 총 회수금: {summary.get('total_returns', 0):,}원")
    print(f"    - 총 수익금: {summary.get('total_profit', 0):,}원 ({summary.get('total_profit_rate', 0)}%)")
    print(f"    - 평균 수익률: {summary.get('avg_return', 0)}%")
    print(f"    - 수익 종목: {summary.get('success_count', 0)}개")
    print(f"    - 손실 종목: {summary.get('fail_count', 0)}개")

    return yesterday_df, summary


if __name__ == "__main__":
    # 테스트
    yesterday_df, summary = update_with_next_day_results()
    if yesterday_df is not None:
        print(yesterday_df.head(10))
